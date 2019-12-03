##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2019 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from django.test import TestCase
from django.utils.translation import gettext_lazy as _

from base.models.enums.prerequisite_operator import AND, OR
from base.tests.factories.education_group_year import EducationGroupYearFactory
from base.tests.factories.group_element_year import GroupElementYearChildLeafFactory, GroupElementYearFactory
from base.tests.factories.person import PersonFactory
from base.tests.factories.prerequisite import PrerequisiteFactory
from program_management.business.excel import EducationGroupYearLearningUnitsPrerequisitesToExcel, \
    EducationGroupYearLearningUnitsIsPrerequisiteOfToExcel, _get_blocks_prerequisite_of, FIX_TITLES, _get_headers, \
    optional_header_for_proposition, optional_header_for_credits, optional_header_for_volume, _get_attribution_line, \
    _fix_data, _get_workbook_for_custom_xls, _build_legend_sheet, LEGEND_WB_CONTENT, LEGEND_WB_STYLE, _optional_data,\
    _build_excel_lines_ues
from program_management.forms.custom_xls import CustomXlsForm
from base.business.learning_unit_xls import CREATION_COLOR, MODIFICATION_COLOR, TRANSFORMATION_COLOR, \
    TRANSFORMATION_AND_MODIFICATION_COLOR, SUPPRESSION_COLOR
from openpyxl.styles import Style, Font
from program_management.business.excel import EducationGroupYearLearningUnitsContainedToExcel


class TestGeneratePrerequisitesWorkbook(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.education_group_year = EducationGroupYearFactory()
        cls.child_leaves = GroupElementYearChildLeafFactory.create_batch(
            6,
            parent=cls.education_group_year
        )
        luy_acronyms = ["LCORS124" + str(i) for i in range(0, len(cls.child_leaves))]
        for node, acronym in zip(cls.child_leaves, luy_acronyms):
            node.child_leaf.acronym = acronym
            node.child_leaf.save()

        cls.luy_children = [child.child_leaf for child in cls.child_leaves]

        PrerequisiteFactory(
            learning_unit_year=cls.luy_children[0],
            education_group_year=cls.education_group_year,
            items__groups=(
                (cls.luy_children[1],),
            )
        )
        PrerequisiteFactory(
            learning_unit_year=cls.luy_children[2],
            education_group_year=cls.education_group_year,
            items__groups=(
                (cls.luy_children[3],),
                (cls.luy_children[4], cls.luy_children[5])
            )
        )
        cls.workbook_prerequisites = \
            EducationGroupYearLearningUnitsPrerequisitesToExcel(cls.education_group_year)._to_workbook()
        cls.workbook_is_prerequisite = \
            EducationGroupYearLearningUnitsIsPrerequisiteOfToExcel(cls.education_group_year)._to_workbook()
        cls.sheet_prerequisites = cls.workbook_prerequisites.worksheets[0]
        cls.sheet_is_prerequisite = cls.workbook_is_prerequisite.worksheets[0]

    def test_header_lines(self):
        expected_headers = [
            [self.education_group_year.acronym, self.education_group_year.title, _('Code'), _('Title'),
             _('Cred. rel./abs.'), _('Block'), _('Mandatory')],
            [_("Official"), None, None, None, None, None, None]
        ]

        headers = [row_to_value(row) for row in self.sheet_prerequisites.iter_rows(range_string="A1:G2")]
        self.assertListEqual(headers, expected_headers)

    def test_when_learning_unit_year_has_one_prerequisite(self):
        expected_content = [
            [self.luy_children[0].acronym, self.luy_children[0].complete_title, None, None, None, None, None],

            [_("has as prerequisite") + " :", '',
             self.luy_children[1].acronym,
             self.luy_children[1].complete_title_i18n,
             "{} / {}".format(self.child_leaves[1].relative_credits, self.luy_children[1].credits),
             str(self.child_leaves[1].block) if self.child_leaves[1].block else '',
             _("Yes") if self.child_leaves[1].is_mandatory else _("No")]
        ]

        content = [row_to_value(row) for row in self.sheet_prerequisites.iter_rows(range_string="A3:G4")]
        self.assertListEqual(expected_content, content)

    def test_when_learning_unit_year_has_multiple_prerequisites(self):
        expected_content = [
            [self.luy_children[2].acronym, self.luy_children[2].complete_title, None, None, None, None, None],

            [_("has as prerequisite") + " :", '', self.luy_children[3].acronym,
             self.luy_children[3].complete_title_i18n,
             "{} / {}".format(self.child_leaves[3].relative_credits, self.luy_children[3].credits),
             str(self.child_leaves[3].block) if self.child_leaves[3].block else '',
             _("Yes") if self.child_leaves[3].is_mandatory else _("No")],

            ['', _(AND), "(" + self.luy_children[4].acronym, self.luy_children[4].complete_title_i18n,
             "{} / {}".format(self.child_leaves[4].relative_credits, self.luy_children[4].credits),
             str(self.child_leaves[4].block) if self.child_leaves[4].block else '',
             _("Yes") if self.child_leaves[4].is_mandatory else _("No")
             ],

            ['', _(OR), self.luy_children[5].acronym + ")", self.luy_children[5].complete_title_i18n,
             "{} / {}".format(self.child_leaves[5].relative_credits, self.luy_children[5].credits),
             str(self.child_leaves[5].block) if self.child_leaves[5].block else '',
             _("Yes") if self.child_leaves[5].is_mandatory else _("No")
             ]
        ]
        content = [row_to_value(row) for row in self.sheet_prerequisites.iter_rows(range_string="A5:G8")]
        self.assertListEqual(expected_content, content)

    def test_get_blocks_prerequisite_of(self):
        gey = GroupElementYearFactory(block=123)
        self.assertEqual(_get_blocks_prerequisite_of(gey), '1 ; 2 ; 3')
        gey = GroupElementYearFactory(block=1)
        self.assertEqual(_get_blocks_prerequisite_of(gey), '1')


def row_to_value(sheet_row):
    return [cell.value for cell in sheet_row]


class TestGenerateEducationGroupYearLearningUnitsContainedWorkbook(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.education_group_year = EducationGroupYearFactory()
        cls.child_leaves = GroupElementYearChildLeafFactory.create_batch(
            2,
            parent=cls.education_group_year,
            is_mandatory=True
        )
        for node, acronym in zip(cls.child_leaves, ["LCORS124" + str(i) for i in range(0, len(cls.child_leaves))]):
            node.child_leaf.acronym = acronym
            node.child_leaf.save()
        cls.luy_children = [child.child_leaf for child in cls.child_leaves]

    def test_header_lines_without_optional_titles(self):
        custom_xls_form = CustomXlsForm({})
        expected_headers = FIX_TITLES

        self.assertListEqual(_get_headers(custom_xls_form)[0], expected_headers)

    def test_header_lines_with_optional_titles(self):
        custom_xls_form = CustomXlsForm({'proposition': 'on',
                                         'credits': 'on',
                                         'volume': 'on'})

        expected_headers = \
            FIX_TITLES + optional_header_for_proposition + optional_header_for_credits + optional_header_for_volume
        self.assertListEqual(_get_headers(custom_xls_form)[0], expected_headers)

    def test_get_attribution_line(self):
        person = PersonFactory(last_name='Last', first_name='First', middle_name='Middle')
        self.assertEqual(_get_attribution_line(person), 'LAST First Middle')

    def test_fix_data(self):
        gey = self.child_leaves[0]
        luy = self.luy_children[0]
        expected = [luy.acronym,
                    luy.academic_year.year,
                    luy.complete_title_i18n,
                    luy.get_container_type_display(),
                    luy.get_subtype_display(),
                    "{} - {}".format(gey.parent.partial_acronym, gey.parent.title),
                    "{} / {}".format(gey.relative_credits or '-', luy.credits.normalize() or '-'),
                    gey.block or '',
                    _('yes')
                    ]
        res = _fix_data(gey, luy)
        self.assertEqual(res, expected)

    def test_legend_workbook_exists(self):
        wb = _get_workbook_for_custom_xls([['header'], [['row1 col1']]], True)
        self.assertEqual(len(wb.worksheets), 2)

    def test_legend_workbook_do_not_exists(self):
        wb = _get_workbook_for_custom_xls([['header'], [['row1 col1']]], False)
        self.assertEqual(len(wb.worksheets), 1)

    def test_legend_workbook_content(self):
        expected_content = [[_("Creation")], [_("Modification")], [_("Transformation")],
                            [_("Transformation and modification")], [_("Suppression")]]

        data = _build_legend_sheet()
        self.assertListEqual(data.get(LEGEND_WB_CONTENT), expected_content)

    def test_legend_workbook_style(self):
        data = _build_legend_sheet()
        self.assertListEqual(data.get(LEGEND_WB_STYLE).get(Style(font=Font(color=CREATION_COLOR))), [1])
        self.assertListEqual(data.get(LEGEND_WB_STYLE).get(Style(font=Font(color=MODIFICATION_COLOR))), [2])
        self.assertListEqual(data.get(LEGEND_WB_STYLE).get(Style(font=Font(color=TRANSFORMATION_COLOR))), [3])
        self.assertListEqual(
            data.get(LEGEND_WB_STYLE).get(Style(font=Font(color=TRANSFORMATION_AND_MODIFICATION_COLOR))),
            [4])
        self.assertListEqual(data.get(LEGEND_WB_STYLE).get(Style(font=Font(color=SUPPRESSION_COLOR))), [5])

    def test_no_optional_data_to_add(self):
        form = CustomXlsForm({})
        self.assertDictEqual(_optional_data(form),
                             {'has_required_entity': False,
                              'has_proposition': False,
                              'has_credits': False,
                              'has_allocation_entity': False,
                              'has_english_title': False,
                              'has_teacher_list': False,
                              'has_periodicity': False,
                              'has_active': False,
                              'has_volume': False,
                              'has_quadrimester': False,
                              'has_session_derogation': False,
                              'has_language': False,
                              }
                             )

    def test_all_optional_data_to_add(self):
        form = CustomXlsForm({'required_entity': 'on',
                              'proposition': 'on',
                              'credits': 'on',
                              'allocation_entity': 'on',
                              'english_title': 'on',
                              'teacher_list': 'on',
                              'periodicity': 'on',
                              'active': 'on',
                              'volume': 'on',
                              'quadrimester': 'on',
                              'session_derogation': 'on',
                              'language': 'on'})
        self.assertDictEqual(_optional_data(form),
                             {'has_required_entity': True,
                              'has_proposition': True,
                              'has_credits': True,
                              'has_allocation_entity': True,
                              'has_english_title': True,
                              'has_teacher_list': True,
                              'has_periodicity': True,
                              'has_active': True,
                              'has_volume': True,
                              'has_quadrimester': True,
                              'has_session_derogation': True,
                              'has_language': True,
                              }
                             )

    def test_data(self):
        custom_form = CustomXlsForm({})
        exl = EducationGroupYearLearningUnitsContainedToExcel(self.education_group_year, custom_form)
        excel_lines = _build_excel_lines_ues(exl.egy, custom_form, exl.learning_unit_years_parent)

        idx = 1

        for k, gey in exl.learning_unit_years_parent.items():
            luy = gey.child_leaf
            expected = [luy.acronym,
                        luy.academic_year.year,
                        luy.complete_title_i18n,
                        luy.get_container_type_display(),
                        luy.get_subtype_display(),
                        "{} - {}".format(gey.parent.partial_acronym, gey.parent.title),
                        "{} / {}".format(gey.relative_credits or '-', luy.credits.normalize() or '-'),
                        gey.block or '',
                        _('yes')
                        ]
            self.assertListEqual(excel_lines[idx], expected)
            idx += 1
