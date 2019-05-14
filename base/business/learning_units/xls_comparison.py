##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2019 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from django.utils.translation import ugettext_lazy as _
from openpyxl.utils import get_column_letter

from base.business import learning_unit_year_with_context
from base.business.entity import build_entity_container_prefetch
from base.business.learning_unit import get_organization_from_learning_unit_year
from base.business.learning_unit_year_with_context import append_latest_entities, append_components, \
    get_learning_component_prefetch
from base.business.learning_units.comparison import get_partims_as_str
from base.business.proposal_xls import BLANK_VALUE, XLS_DESCRIPTION_COMPARISON, XLS_COMPARISON_FILENAME, \
    COMPARISON_PROPOSAL_TITLES, COMPARISON_WORKSHEET_TITLE, basic_titles, components_titles
from base.business.utils.convert import volume_format
from base.business.xls import get_name_or_username
from base.models.academic_year import current_academic_year
from base.models.campus import find_by_id as find_campus_by_id
from base.models.entity import find_by_id
from base.models.enums import entity_container_year_link_type as entity_types, vacant_declaration_type, \
    attribution_procedure
from base.models.enums import learning_component_year_type
from base.models.enums.component_type import DEFAULT_ACRONYM_COMPONENT
from base.models.enums.learning_component_year_type import LECTURING, PRACTICAL_EXERCISES
from base.models.enums.learning_container_year_types import LearningContainerYearType
from base.models.enums.learning_unit_year_periodicity import PERIODICITY_TYPES
from base.models.external_learning_unit_year import ExternalLearningUnitYear
from base.models.learning_unit_year import LearningUnitYear, get_by_id
from base.models.proposal_learning_unit import ProposalLearningUnit
from osis_common.document import xls_build
from reference.models.language import find_by_id as find_language_by_id

EMPTY_VALUE = ''
DATE_FORMAT = '%d-%m-%Y'
DATE_TIME_FORMAT = '%d-%m-%Y %H:%M'
DESC = "desc"
WORKSHEET_TITLE = 'learning_units_comparison'
XLS_FILENAME = 'learning_units_comparison'
XLS_DESCRIPTION = _("Comparison of learning units")

ACRONYM_COL_NUMBER = 0
ACADEMIC_COL_NUMBER = 1
CELLS_MODIFIED_NO_BORDER = 'modifications'
CELLS_TOP_BORDER = 'border_not_modified'
DATA = 'data'


def learning_unit_titles():
    return basic_titles() + components_titles()


def create_xls_comparison(user, learning_unit_years, filters, academic_yr_comparison):
    working_sheets_data = []
    cells_modified_with_green_font = []
    cells_with_top_border = []

    if learning_unit_years:
        luys_for_2_years = _get_learning_unit_yrs_on_2_different_years(academic_yr_comparison, learning_unit_years)
        data = prepare_xls_content(luys_for_2_years)
        working_sheets_data = data.get('data')
        cells_modified_with_green_font = data.get(CELLS_MODIFIED_NO_BORDER)
        cells_with_top_border = data.get(CELLS_TOP_BORDER)
    parameters = {
        xls_build.DESCRIPTION: XLS_DESCRIPTION,
        xls_build.USER: get_name_or_username(user),
        xls_build.FILENAME: XLS_FILENAME,
        xls_build.HEADER_TITLES: learning_unit_titles(),
        xls_build.WS_TITLE: WORKSHEET_TITLE,
    }
    dict_styled_cells = {}
    if cells_modified_with_green_font:
        dict_styled_cells[xls_build.STYLE_MODIFIED] = cells_modified_with_green_font

    if cells_with_top_border:
        dict_styled_cells[xls_build.STYLE_BORDER_TOP] = cells_with_top_border

    if dict_styled_cells:
        parameters[xls_build.STYLED_CELLS] = dict_styled_cells

    return xls_build.generate_xls(xls_build.prepare_xls_parameters_list(working_sheets_data, parameters), filters)


def _get_learning_unit_yrs_on_2_different_years(academic_yr_comparison, learning_unit_years):
    learning_unit_years = LearningUnitYear.objects.filter(
        learning_unit__in=(_get_learning_units(learning_unit_years)),
        academic_year__year__in=(
            learning_unit_years[0].academic_year.year,
            academic_yr_comparison)
    ).select_related(
        'academic_year',
        'learning_container_year',
        'learning_container_year__academic_year'
    ).prefetch_related(
        get_learning_component_prefetch()
    ).prefetch_related(
        build_entity_container_prefetch([
            entity_types.ALLOCATION_ENTITY,
            entity_types.REQUIREMENT_ENTITY,
            entity_types.ADDITIONAL_REQUIREMENT_ENTITY_1,
            entity_types.ADDITIONAL_REQUIREMENT_ENTITY_2
        ])
    ).order_by('learning_unit', 'academic_year__year')
    [append_latest_entities(learning_unit) for learning_unit in learning_unit_years]
    [append_components(learning_unit) for learning_unit in learning_unit_years]
    return learning_unit_years


def _get_learning_units(learning_unit_years):
    return list(set([l.learning_unit for l in learning_unit_years]))


def prepare_xls_content(learning_unit_yrs):
    data = []
    learning_unit = None
    first_data = None
    modified_cells_no_border = []
    top_border = []
    for line_index, l_u_yr in enumerate(learning_unit_yrs, start=1):

        if learning_unit is None:
            learning_unit = l_u_yr.learning_unit
            new_line = True
        else:
            if learning_unit == l_u_yr.learning_unit:
                new_line = False
            else:
                learning_unit = l_u_yr.learning_unit
                new_line = True
        luy_data = extract_xls_data_from_learning_unit(l_u_yr, new_line, first_data)
        if new_line:
            first_data = luy_data
            top_border.extend(get_border_columns(line_index + 1))
        else:
            modified_cells_no_border.extend(
                _check_changes_other_than_code_and_year(first_data, luy_data, line_index + 1))
            first_data = None
        data.append(luy_data)

    return {
        DATA: data,
        CELLS_TOP_BORDER: top_border or None,
        CELLS_MODIFIED_NO_BORDER: modified_cells_no_border or None,
    }


def extract_xls_data_from_learning_unit(learning_unit_yr, new_line, first_data):
    data = _get_data(learning_unit_yr, new_line, first_data)
    data.extend(_component_data(learning_unit_yr.components, learning_component_year_type.LECTURING))
    data.extend(_component_data(learning_unit_yr.components, learning_component_year_type.PRACTICAL_EXERCISES))
    return data


def translate_status(value):
    if value:
        return _('Active')
    else:
        return _('Inactive')


def _component_data(components, learning_component_yr_type):
    if components:
        for component in components:
            if component.type == learning_component_yr_type:
                return _get_volumes(component, components)
    return [EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE, EMPTY_VALUE,
            EMPTY_VALUE, EMPTY_VALUE]


def _get_data(learning_unit_yr, new_line, first_data, partims=True):
    organization = get_organization_from_learning_unit_year(learning_unit_yr)

    data = [
        _get_acronym(learning_unit_yr, new_line, first_data),
        learning_unit_yr.academic_year.name,
        learning_unit_yr.learning_container_year.get_container_type_display()
        if learning_unit_yr.learning_container_year.container_type else EMPTY_VALUE,
        translate_status(learning_unit_yr.status),
        learning_unit_yr.get_subtype_display() if learning_unit_yr.subtype else EMPTY_VALUE,
        learning_unit_yr.get_internship_subtype_display() if learning_unit_yr.internship_subtype else EMPTY_VALUE,
        volume_format(learning_unit_yr.credits),
        learning_unit_yr.language.name if learning_unit_yr.language else EMPTY_VALUE,
        learning_unit_yr.get_periodicity_display() if learning_unit_yr.periodicity else EMPTY_VALUE,
        get_translation(learning_unit_yr.quadrimester),
        get_translation(learning_unit_yr.session),
        learning_unit_yr.learning_container_year.common_title,
        learning_unit_yr.specific_title,
        learning_unit_yr.learning_container_year.common_title_english,
        learning_unit_yr.specific_title_english,
        _get_entity_to_display(learning_unit_yr.entities.get(entity_types.REQUIREMENT_ENTITY)),
        _get_entity_to_display(learning_unit_yr.entities.get(entity_types.ALLOCATION_ENTITY)),
        _get_entity_to_display(learning_unit_yr.entities.get(entity_types.ADDITIONAL_REQUIREMENT_ENTITY_1)),
        _get_entity_to_display(learning_unit_yr.entities.get(entity_types.ADDITIONAL_REQUIREMENT_ENTITY_2)),
        _('Yes') if learning_unit_yr.professional_integration else _('No'),
        organization.name if organization else EMPTY_VALUE,
        learning_unit_yr.campus if learning_unit_yr.campus else EMPTY_VALUE]
    if partims:
        data.append(get_partims_as_str(learning_unit_yr.get_partims_related()))
    data.extend(
        [
            learning_unit_yr.learning_unit.faculty_remark,
            learning_unit_yr.learning_unit.other_remark,
            _('Yes') if learning_unit_yr.learning_container_year.team else _('No'),
            _('Yes') if learning_unit_yr.learning_container_year.is_vacant else _('No'),
            learning_unit_yr.learning_container_year.get_type_declaration_vacant_display(),
            learning_unit_yr.get_attribution_procedure_display(),
        ]
    )

    return data


def _get_acronym(learning_unit_yr, new_line, first_data):
    if first_data:
        acronym = EMPTY_VALUE
        if new_line:
            acronym = learning_unit_yr.acronym
        else:
            if learning_unit_yr.acronym != first_data[ACRONYM_COL_NUMBER]:
                acronym = learning_unit_yr.acronym
        return acronym
    else:
        return learning_unit_yr.acronym


def _get_volumes(component, components):
    volumes = components[component]
    return [
        component.acronym if component.acronym else EMPTY_VALUE,
        volumes.get('VOLUME_Q1', EMPTY_VALUE),
        volumes.get('VOLUME_Q2', EMPTY_VALUE),
        volumes.get('VOLUME_TOTAL', EMPTY_VALUE),
        component.real_classes if component.real_classes else EMPTY_VALUE,
        component.planned_classes if component.planned_classes else EMPTY_VALUE,
        volumes.get('VOLUME_GLOBAL', '0'),
        volumes.get('VOLUME_REQUIREMENT_ENTITY', EMPTY_VALUE),
        volumes.get('VOLUME_ADDITIONAL_REQUIREMENT_ENTITY_1', EMPTY_VALUE),
        volumes.get('VOLUME_ADDITIONAL_REQUIREMENT_ENTITY_2', EMPTY_VALUE)
    ]


def get_translation(value):
    return str(_(value)) if value else EMPTY_VALUE


def _get_entity_to_display(entity):
    return entity.acronym if entity else EMPTY_VALUE


def get_academic_year_of_reference(objects):
    """ TODO : Has to be improved because it's not optimum if the xls list is created from a search with a
    criteria : academic_year = 'ALL' """
    if objects:
        return _get_academic_year(objects[0])
    return current_academic_year()


def _get_academic_year(obj):
    if isinstance(obj, LearningUnitYear):
        return obj.academic_year
    if isinstance(obj, (ProposalLearningUnit, ExternalLearningUnitYear)):
        return obj.learning_unit_year.academic_year


def _check_changes_other_than_code_and_year(first_data, second_data, line_index):
    modifications = []
    for col_index, obj in enumerate(first_data):
        if col_index == ACRONYM_COL_NUMBER and second_data[ACRONYM_COL_NUMBER] != EMPTY_VALUE:
            modifications.append('{}{}'.format(get_column_letter(col_index + 1), line_index))
        else:
            if obj != second_data[col_index] and col_index != ACADEMIC_COL_NUMBER:
                modifications.append('{}{}'.format(get_column_letter(col_index + 1), line_index))

    return modifications


def get_border_columns(line):
    style = []
    for col_index, obj in enumerate(learning_unit_titles(), start=1):
        style.append('{}{}'.format(get_column_letter(col_index), line))
    return style


def _get_component_data_by_type(component, type):
    if component:
        return [
            DEFAULT_ACRONYM_COMPONENT.get(type),
            component.get('VOLUME_Q1'),
            component.get('VOLUME_Q2'),
            component.get('VOLUME_TOTAL'),
            component.get('REAL_CLASSES'),
            component.get('PLANNED_CLASSES'),
            component.get('VOLUME_TOTAL_REQUIREMENT_ENTITIES'),
            component.get('VOLUME_REQUIREMENT_ENTITY'),
            component.get('VOLUME_ADDITIONAL_REQUIREMENT_ENTITY_1'),
            component.get('VOLUME_ADDITIONAL_REQUIREMENT_ENTITY_2'),
        ]
    else:
        return []


def _get_learning_unit_yr_with_component(learning_unit_years):
    learning_unit_years = LearningUnitYear.objects.filter(
        learning_unit_year_id__in=[luy.id for luy in learning_unit_years]
        ).select_related(
        'academic_year',
        'learning_container_year',
        'learning_container_year__academic_year'
    ).prefetch_related(
        get_learning_component_prefetch()
    ).prefetch_related(
        build_entity_container_prefetch([
            entity_types.ALLOCATION_ENTITY,
            entity_types.REQUIREMENT_ENTITY,
            entity_types.ADDITIONAL_REQUIREMENT_ENTITY_1,
            entity_types.ADDITIONAL_REQUIREMENT_ENTITY_2
        ])
    ).order_by('learning_unit', 'academic_year__year')
    [append_latest_entities(learning_unit) for learning_unit in learning_unit_years]
    [append_components(learning_unit) for learning_unit in learning_unit_years]
    return learning_unit_years


def prepare_xls_content_for_comparison(luy_with_proposals):
    line_index = 1
    data = []
    top_border = []
    modified_cells_no_border = []
    for luy_with_proposal in luy_with_proposals:
        top_border.extend(get_border_columns(line_index + 1))
        data_proposal = _get_proposal_data(luy_with_proposal)
        data.append(data_proposal)

        proposal = luy_with_proposal.proposallearningunit
        initial_luy_data = proposal.initial_data

        if initial_luy_data:
            initial_data = _get_data_from_initial_data(luy_with_proposal.proposallearningunit.initial_data)
            data.append(initial_data)
        else:
            initial_data = []

        modified_cells_no_border.extend(
            _check_changes(initial_data,
                           data_proposal,
                           line_index + 2))
        line_index = line_index + 2

    return {
        DATA: data,
        CELLS_TOP_BORDER: top_border or None,
        CELLS_MODIFIED_NO_BORDER: modified_cells_no_border or None,
    }


def _get_data_from_initial_data(initial_data):
    learning_unit_yr = get_by_id(initial_data.get('learning_unit_year')['id'])
    requirement_entity = find_by_id(initial_data.get('entities')['REQUIREMENT_ENTITY'])
    allocation_entity = find_by_id(initial_data.get('entities')['ALLOCATION_ENTITY'])
    add1_requirement_entity = find_by_id(initial_data.get('entities')['ADDITIONAL_REQUIREMENT_ENTITY_1'])
    add2_requirement_entity = find_by_id(initial_data.get('entities')['ADDITIONAL_REQUIREMENT_ENTITY_2'])
    campus = find_campus_by_id(initial_data.get('learning_unit_year')['campus'])

    organization = get_organization_from_learning_unit_year(learning_unit_yr)
    language = find_language_by_id(initial_data.get('learning_unit_year')['language'])
    lu_initial = initial_data.get('learning_unit', None)
    luy_initial = initial_data.get('learning_unit_year', None)
    lcy_initial = initial_data.get('learning_container_year', None)

    data = [
        str(_('Initial data')),
        luy_initial['acronym'],
        learning_unit_yr.academic_year.name,
        dict(LearningContainerYearType.choices())[lcy_initial['container_type']] if
        lcy_initial['container_type'] else '-',
        translate_status(luy_initial['status']),
        learning_unit_yr.get_subtype_display(),
        get_translation(luy_initial['internship_subtype']),
        volume_format(luy_initial['credits']),
        language.name if language else EMPTY_VALUE,
        dict(PERIODICITY_TYPES)[luy_initial['periodicity']] if luy_initial['periodicity'] else BLANK_VALUE,
        get_translation(luy_initial['quadrimester']),
        get_translation(luy_initial['session']),
        get_representing_string(lcy_initial['common_title']),
        get_representing_string(luy_initial['specific_title']),
        get_representing_string(lcy_initial['common_title_english']),
        get_representing_string(luy_initial['specific_title_english']),
        requirement_entity.most_recent_acronym if requirement_entity else BLANK_VALUE,
        allocation_entity.most_recent_acronym if allocation_entity else BLANK_VALUE,
        add1_requirement_entity.most_recent_acronym if add1_requirement_entity else BLANK_VALUE,
        add2_requirement_entity.most_recent_acronym if add2_requirement_entity else BLANK_VALUE,
        _('Yes') if luy_initial['professional_integration'] else _('No'),
        organization.name if organization else BLANK_VALUE,
        campus if campus else BLANK_VALUE,
        get_representing_string(lu_initial['faculty_remark']),
        get_representing_string(lu_initial['other_remark']),
        _('Yes') if lcy_initial.get('team') else _('No'),
        _('Yes') if lcy_initial.get('is_vacant') else _('No'),
        dict(vacant_declaration_type.DECLARATION_TYPE)[lcy_initial.get('type_declaration_vacant')] if lcy_initial.get(
            'type_declaration_vacant') else BLANK_VALUE,
        dict(attribution_procedure.ATTRIBUTION_PROCEDURES)[luy_initial.get('attribution_procedure')] if luy_initial.get(
            'attribution_procedure') else BLANK_VALUE,
    ]
    return _get_data_from_components_initial_data(data, initial_data)


def _check_changes(initial_data, proposal_data, line_index):
    modifications = []
    for col_index, obj in enumerate(initial_data[2:]):
        if obj != proposal_data[col_index+2]:
            modifications.append('{}{}'.format(get_column_letter(col_index+2 + 1), line_index))
    return modifications


def get_representing_string(value):
    return value or BLANK_VALUE


def create_xls_proposal_comparison(user, learning_units_with_proposal, filters):
    data = prepare_xls_content_for_comparison(learning_units_with_proposal)

    working_sheets_data = data.get('data')
    cells_modified_with_green_font = data.get(CELLS_MODIFIED_NO_BORDER)
    cells_with_top_border = data.get(CELLS_TOP_BORDER)

    parameters = {
        xls_build.DESCRIPTION: XLS_DESCRIPTION_COMPARISON,
        xls_build.USER: get_name_or_username(user),
        xls_build.FILENAME: XLS_COMPARISON_FILENAME,
        xls_build.HEADER_TITLES: COMPARISON_PROPOSAL_TITLES,
        xls_build.WS_TITLE: COMPARISON_WORKSHEET_TITLE,
    }
    dict_styled_cells = {}
    if cells_modified_with_green_font:
        dict_styled_cells[xls_build.STYLE_MODIFIED] = cells_modified_with_green_font

    if cells_with_top_border:
        dict_styled_cells[xls_build.STYLE_BORDER_TOP] = cells_with_top_border
    if dict_styled_cells:
        parameters[xls_build.STYLED_CELLS] = dict_styled_cells
    return xls_build.generate_xls(xls_build.prepare_xls_parameters_list(working_sheets_data, parameters), filters)


def _get_basic_components(learning_unit_yr):
    learning_unit_yr = find_learning_unit_yr_with_components_data(learning_unit_yr)
    components = []
    components_values = []
    for key, value in learning_unit_yr.components.items():
        components.append(key)
        components_values.append(value)

    practical_component = None
    lecturing_component = None
    for index, component in enumerate(components):
        if not practical_component and component.type == PRACTICAL_EXERCISES:
            practical_component = _build_component(component.real_classes, components_values, index)

        if not lecturing_component and component.type == LECTURING:
            lecturing_component = _build_component(component.real_classes, components_values, index)
    return {PRACTICAL_EXERCISES: practical_component, LECTURING: lecturing_component}


def _build_component(real_classes, components_values, index):
    a_component = components_values[index]
    a_component['REAL_CLASSES'] = real_classes
    return a_component


def _get_components_data(learning_unit_yr):
    components_data_dict = _get_basic_components(learning_unit_yr)
    return \
        _get_component_data_by_type(components_data_dict.get(LECTURING), LECTURING) + \
        _get_component_data_by_type(components_data_dict.get(PRACTICAL_EXERCISES), PRACTICAL_EXERCISES)


def _get_proposal_data(learning_unit_yr):
    data_proposal = [_('Proposal')] + _get_data(learning_unit_yr, False, None, False)
    data_proposal.extend(_get_components_data(learning_unit_yr))
    return data_proposal


def find_learning_unit_yr_with_components_data(learning_unit_yr):
    learning_unit_yrs = learning_unit_year_with_context.get_with_context(
        learning_container_year_id=learning_unit_yr.learning_container_year.id
    )
    if learning_unit_yrs:
        learning_unit_yr = next(luy for luy in learning_unit_yrs if luy.id == learning_unit_yr.id)

    return learning_unit_yr


def _get_data_from_components_initial_data(data_without_components, initial_data):
    data = data_without_components
    volumes = initial_data.get('volumes')
    if volumes:
        data = data + _get_component_data_by_type(volumes.get(LECTURING), LECTURING)
        data = data + _get_component_data_by_type(volumes.get(PRACTICAL_EXERCISES), PRACTICAL_EXERCISES)
    return data
